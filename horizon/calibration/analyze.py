# SPDX-FileCopyrightText: 2026 Fonden Mærsk Mc-Kinney Møller Center for Zero Carbon Shipping
# SPDX-License-Identifier: Apache-2.0

"""Calibration analysis dashboard.

Reads completed calibration simulation results and generates an interactive
HTML dashboard for comparing parameter values side-by-side across regulation
scenarios.
"""

import concurrent.futures
import csv
import glob
import logging
import os
import warnings
from collections import Counter, defaultdict, namedtuple

import numpy as np
import openpyxl
import pandas as pd
from scipy.stats import spearmanr

logger = logging.getLogger(__name__)

RunInfo = namedtuple("RunInfo", ["folder_name", "regulation", "param_token", "param_value"])
Recommendation = namedtuple(
    "Recommendation",
    ["param_token", "recommended_value", "confidence", "scores_breakdown", "current_default"],
)

# -- Target metrics from the Global sheet --
# Single-column metrics (metric_name -> row1 value)
SINGLE_METRICS = {
    "TotalEquivalentWTW": "Total WTW Emissions",
    "Expenses": "Total Expenses",
}
# Multi-column metrics (metric_name -> display label; sub-columns are fuel types)
MULTI_METRICS = {
    "ConsumedEnergy": "Consumed Energy by Fuel",
    "InstalledPower": "Installed Power by Fuel Type",
}

# Fixed palette for plot lines (5 values per parameter + default highlight)
_LINE_COLORS = [
    "#2c4068",  # blue-7
    "#3c5e86",  # blue-6
    "#68a4c2",  # blue-5
    "#96c8e4",  # blue-4
    "#b8e4f4",  # blue-3
]
_DEFAULT_COLOR = "#a2703c"  # yellow-7

# Fuel types to track in share plots
_FUEL_TYPES = ["OIL", "METHANE", "METHANOL", "AMMONIA"]

# Per-fuel colors for share plots (one distinct color per fuel)
_FUEL_COLORS = {
    "OIL": "#323232",       # black-7
    "METHANE": "#3c5e86",   # blue-6
    "METHANOL": "#286464",  # green-6
    "AMMONIA": "#a2703c",   # yellow-7
}

# Expected no-regulation 2050 fuel share ranges (midpoint, half-width)
_NO_REG_SHARE_BENCHMARKS = {
    "OIL": (0.50, 0.05),        # 45-55%
    "METHANE": (0.375, 0.025),   # 35-40%
    "METHANOL": (0.05, 0.05),    # 0-10%
    "AMMONIA": (0.01, 0.01),     # 0-2%
}

# Scoring weights
_SCORE_WEIGHTS = {
    "sensitivity": 0.25,
    "monotonicity": 0.15,
    "cross_scenario": 0.20,
    "fuel_share_realism": 0.25,
    "smoothness": 0.15,
}


# ---------------------------------------------------------------------------
# 1. Load index
# ---------------------------------------------------------------------------

def _load_index(calibration_dir):
    """Read sampled_parameters.csv and return a list of RunInfo tuples.

    Parameters
    ----------
    calibration_dir : str
        Root directory containing calibration output folders and the CSV.

    Returns
    -------
    list[RunInfo]
        One entry per simulation folder.
    dict
        Parameter defaults: {param_token: default_value}.
    """
    # Find the CSV — could be at calibration_dir or one level up
    csv_path = None
    for candidate in [
        os.path.join(calibration_dir, "sampled_parameters.csv"),
        *glob.glob(os.path.join(calibration_dir, "**", "sampled_parameters.csv"), recursive=True),
    ]:
        if os.path.isfile(candidate):
            csv_path = candidate
            break

    if csv_path is None:
        raise FileNotFoundError(
            f"Cannot find sampled_parameters.csv in or under {calibration_dir}"
        )

    logger.info("Reading index from %s", csv_path)

    with open(csv_path, newline="") as f:
        reader = csv.reader(f)
        header = next(reader)
        next(reader)  # skip parameter_type row

        # Identify columns
        col_map = {name: idx for idx, name in enumerate(header)}

        # Resolve sample name column (prefer "sample", fall back to "sample_number")
        sample_col = col_map.get("sample", col_map.get("sample_number", 0))

        # Find the regulation column (scenario parameter)
        regulation_col = col_map.get("regulation")

        # Find parameter columns (everything except sample_number, regulation, sample)
        skip_cols = {"sample_number", "sample", "regulation"}
        param_tokens = [h for h in header if h not in skip_cols and h]

        # Read all rows
        rows = list(reader)

    # Detect defaults: for each parameter, find the value that appears when
    # that parameter is NOT being varied. In calibration mode, the sample name
    # is "{param_token}_{value}" — so when the sample doesn't start with a
    # param token, that param is at its default.
    param_defaults = {}
    for pt in param_tokens:
        defaults_seen = set()
        for row in rows:
            sample_name = row[sample_col]
            # If this row is varying a DIFFERENT parameter, the value of pt is its default
            if isinstance(sample_name, str) and not sample_name.startswith(pt + "_"):
                val_str = row[col_map[pt]]
                defaults_seen.add(val_str)
        if len(defaults_seen) == 1:
            param_defaults[pt] = defaults_seen.pop()
        elif defaults_seen:
            # Multiple defaults seen — pick most common
            value_counts = Counter(row[col_map[pt]] for row in rows)
            param_defaults[pt] = max(defaults_seen, key=lambda v: value_counts[v])

    # Detect regulation scenarios.
    # Case 1 (pre-resolved): regulation column exists in CSV — each row has its regulation.
    # Case 2 (legacy): no regulation column — FileHandler expanded scenarios × samples,
    #   producing folders like {regulation}_sample{NNN:03d}. Discover regulations from
    #   folder names in the calibration directory.
    if regulation_col is not None:
        regulations = sorted(set(row[regulation_col] for row in rows))
    else:
        # Scan directory for {regulation}_sample{NNN} folders
        regulations = _discover_regulations(calibration_dir)
        if not regulations:
            regulations = ["default"]

    # Build RunInfo list
    runs = []
    for row in rows:
        sample_name = row[sample_col]
        sample_number = row[col_map.get("sample_number", 0)]

        # Determine which parameter is being varied in this run
        # Calibration sample names are like "bunkering_inertia_0.3"
        varied_param = None
        varied_value = None
        for pt in param_tokens:
            if isinstance(sample_name, str) and sample_name.startswith(pt + "_"):
                varied_param = pt
                varied_value = row[col_map[pt]]
                break

        if varied_param is None:
            logger.warning("Could not identify varied parameter for sample '%s'", sample_name)
            continue

        for regulation in (
            [row[regulation_col]] if regulation_col is not None else regulations
        ):
            # Folder naming depends on whether scenarios were pre-resolved
            if regulation_col is not None:
                # Pre-resolved: sample name IS the suffix (not "sample_N" format)
                folder_name = f"{regulation}_{sample_name}"
            else:
                # Legacy: FileHandler uses sample{NNN:03d} format
                folder_name = f"{regulation}_sample{int(sample_number):03d}"

            runs.append(RunInfo(
                folder_name=folder_name,
                regulation=regulation,
                param_token=varied_param,
                param_value=varied_value,
            ))

    logger.info("Loaded %d calibration runs across %d parameters",
                len(runs), len(set(r.param_token for r in runs)))

    return runs, param_defaults


def _discover_regulations(calibration_dir):
    """Discover regulation scenario names from folder names in the directory.

    Looks for entries matching ``{prefix}_sample{NNN}`` and extracts the prefix
    as the regulation name. Skips the ``os.path.isdir`` check to avoid issues
    on network mounts (SMB/CIFS) where it can return False for directories.
    """
    import re
    pattern = re.compile(r'^(.+)_sample\d{3}$')
    regulations = set()
    for entry in os.listdir(calibration_dir):
        m = pattern.match(entry)
        if m:
            regulations.add(m.group(1))
    return sorted(regulations)


# ---------------------------------------------------------------------------
# 2. Read Excel report
# ---------------------------------------------------------------------------

def _find_report(folder_path):
    """Find the report .xlsx file in a simulation folder.

    Searches in the folder itself and any immediate subdirectory (plots/,
    reports/, etc.) for the first .xlsx file found.
    """
    # Check immediate subdirectories first (plots/, reports/, etc.)
    xlsx_files = glob.glob(os.path.join(folder_path, "*", "*.xlsx"))
    if xlsx_files:
        return xlsx_files[0]
    # Fall back to the folder itself
    xlsx_files = glob.glob(os.path.join(folder_path, "*.xlsx"))
    if xlsx_files:
        return xlsx_files[0]
    return None


def _read_report(folder_path):
    """Read key metrics from a simulation's Excel report (Global sheet).

    Parameters
    ----------
    folder_path : str
        Path to the simulation folder.

    Returns
    -------
    dict or None
        {metric_name: pd.DataFrame} with years as index.
        Returns None if no report found or on error.
    """
    xlsx_path = _find_report(folder_path)
    if xlsx_path is None:
        return None

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        logger.warning("Could not open %s: %s", xlsx_path, e)
        return None

    if "Global" not in wb.sheetnames:
        wb.close()
        return None

    ws = wb["Global"]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(all_rows) < 5:
        return None

    # Row 0: scope, Row 1: metric names, Row 2: sub-metric (fuel type), Row 3: empty, Row 4+: data
    row_metrics = list(all_rows[1])
    row_fuels = list(all_rows[2])

    # Build column mapping: {metric_name: [(col_idx, fuel_label), ...]}
    metric_columns = defaultdict(list)
    for col_idx in range(2, len(row_metrics)):
        metric_name = row_metrics[col_idx]
        if metric_name is None:
            continue
        fuel_label = row_fuels[col_idx] if col_idx < len(row_fuels) else None
        metric_columns[metric_name].append((col_idx, fuel_label))

    # Extract data rows (row 4 onward)
    years = []
    data_rows = []
    for row in all_rows[4:]:
        if row[0] is None:
            continue
        try:
            if hasattr(row[0], "year"):
                years.append(row[0].year)
            else:
                years.append(int(pd.Timestamp(row[0]).year))
        except Exception:
            continue
        data_rows.append(row)

    if not years:
        return None

    result = {}

    # Single-column metrics
    for metric_key in SINGLE_METRICS:
        if metric_key not in metric_columns:
            continue
        cols = metric_columns[metric_key]
        col_idx = cols[0][0]
        values = [_safe_float(row[col_idx]) for row in data_rows]
        result[metric_key] = pd.DataFrame({"Total": values}, index=years)

    # Multi-column metrics
    for metric_key in MULTI_METRICS:
        if metric_key not in metric_columns:
            continue
        cols = metric_columns[metric_key]
        df_data = {}
        for col_idx, fuel_label in cols:
            label = fuel_label or f"col_{col_idx}"
            values = [_safe_float(row[col_idx]) for row in data_rows]
            # Only include if any non-zero values
            if any(v != 0 for v in values):
                df_data[label] = values
        if df_data:
            result[metric_key] = pd.DataFrame(df_data, index=years)

    return result


def _safe_float(v):
    """Convert a cell value to float, returning 0.0 on failure."""
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


# ---------------------------------------------------------------------------
# 3. Build Plotly figures
# ---------------------------------------------------------------------------

def _build_figures(runs, data_by_folder, param_defaults):
    """Build Plotly figures grouped by (regulation, param_token).

    Returns
    -------
    dict
        {(regulation, param_token): list[plotly.graph_objects.Figure]}
    """
    import plotly.graph_objects as go

    # Group runs by (regulation, param_token)
    groups = defaultdict(list)
    for run in runs:
        groups[(run.regulation, run.param_token)].append(run)

    figures = {}

    for (regulation, param_token), group_runs in sorted(groups.items()):
        # Sort by param value
        group_runs = sorted(group_runs, key=lambda r: _safe_float(r.param_value))
        default_val = param_defaults.get(param_token)

        figs = []

        # Build one figure per metric
        all_metrics = list(SINGLE_METRICS.items()) + list(MULTI_METRICS.items())
        for metric_key, metric_label in all_metrics:
            fig = go.Figure()
            has_data = False

            for i, run in enumerate(group_runs):
                report = data_by_folder.get(run.folder_name)
                if report is None or metric_key not in report:
                    continue

                df = report[metric_key]
                is_default = (str(run.param_value) == str(default_val))
                color = _DEFAULT_COLOR if is_default else _LINE_COLORS[i % len(_LINE_COLORS)]
                width = 3 if is_default else 1.5
                label_suffix = " (default)" if is_default else ""

                y_values = df.iloc[:, 0] if metric_key in SINGLE_METRICS else df.sum(axis=1)
                fig.add_trace(go.Scatter(
                    x=df.index.tolist(),
                    y=y_values.tolist(),
                    mode="lines",
                    name=f"{run.param_value}{label_suffix}",
                    line=dict(color=color, width=width),
                ))
                has_data = True

            if has_data:
                fig.update_layout(
                    title=dict(text=f"{metric_label}", font=dict(size=14)),
                    xaxis_title="Year",
                    yaxis_title=metric_label,
                    template="plotly_white",
                    height=350,
                    margin=dict(l=60, r=30, t=50, b=40),
                    legend=dict(
                        orientation="h",
                        yanchor="bottom",
                        y=-0.3,
                        xanchor="center",
                        x=0.5,
                    ),
                )
                figs.append(fig)

        if figs:
            figures[(regulation, param_token)] = figs

    return figures


# ---------------------------------------------------------------------------
# 3b. Fuel share plots
# ---------------------------------------------------------------------------

def _build_fuel_share_figures(runs, data_by_folder, param_defaults):
    """Build per-fuel InstalledPower share plots for each (regulation, param_token).

    Returns
    -------
    dict
        {(regulation, param_token): list[go.Figure]} — one figure per fuel type.
    """
    import plotly.graph_objects as go

    groups = defaultdict(list)
    for run in runs:
        groups[(run.regulation, run.param_token)].append(run)

    figures = {}

    for (regulation, param_token), group_runs in sorted(groups.items()):
        group_runs = sorted(group_runs, key=lambda r: _safe_float(r.param_value))
        default_val = param_defaults.get(param_token)

        figs = []
        for fuel in _FUEL_TYPES:
            fig = go.Figure()
            has_data = False

            for i, run in enumerate(group_runs):
                report = data_by_folder.get(run.folder_name)
                if report is None or "InstalledPower" not in report:
                    continue

                df = report["InstalledPower"]
                fuel_col = _find_fuel_column(df.columns, fuel)
                if fuel_col is None:
                    continue

                row_totals = df.sum(axis=1)
                shares = (df[fuel_col] / row_totals.replace(0, np.nan) * 100).fillna(0)

                is_default = (str(run.param_value) == str(default_val))
                color = _DEFAULT_COLOR if is_default else _LINE_COLORS[i % len(_LINE_COLORS)]
                width = 3 if is_default else 1.5
                label_suffix = " (default)" if is_default else ""

                fig.add_trace(go.Scatter(
                    x=df.index.tolist(),
                    y=shares.tolist(),
                    mode="lines",
                    name=f"{run.param_value}{label_suffix}",
                    line=dict(color=color, width=width),
                ))
                has_data = True

            if has_data:
                fig.update_layout(
                    title=dict(text=f"{fuel} Share of Installed Power", font=dict(size=14)),
                    xaxis_title="Year",
                    yaxis_title="Share (%)",
                    yaxis=dict(range=[0, 100]),
                    template="plotly_white",
                    height=300,
                    margin=dict(l=60, r=30, t=50, b=40),
                    legend=dict(
                        orientation="h", yanchor="bottom", y=-0.35,
                        xanchor="center", x=0.5,
                    ),
                )
                figs.append(fig)

        if figs:
            figures[(regulation, param_token)] = figs

    return figures


def _find_fuel_column(columns, fuel_name):
    """Find a DataFrame column matching a fuel name (case-insensitive partial match)."""
    for col in columns:
        if fuel_name.lower() in col.lower():
            return col
    return None


# ---------------------------------------------------------------------------
# 3c. Scoring functions
# ---------------------------------------------------------------------------

def _get_final_year_values(data_by_value, metric_key="TotalEquivalentWTW"):
    """Extract final-year metric value for each parameter value.

    Parameters
    ----------
    data_by_value : dict
        {param_value_str: report_dict}

    Returns
    -------
    list[tuple[float, float]]
        Sorted (param_value, final_metric_value) pairs.
    """
    pairs = []
    for val_str, report in data_by_value.items():
        if report is None or metric_key not in report:
            continue
        df = report[metric_key]
        if metric_key in SINGLE_METRICS:
            final = df.iloc[-1, 0]
        else:
            final = df.iloc[-1].sum()
        pairs.append((_safe_float(val_str), float(final)))
    return sorted(pairs, key=lambda p: p[0])


def _score_sensitivity(data_by_value, metric_key="TotalEquivalentWTW"):
    """Score how much a parameter affects outcomes (0-1).

    Computes percentage spread in final-year values; 30% spread = max score.
    """
    pairs = _get_final_year_values(data_by_value, metric_key)
    if len(pairs) < 2:
        return 0.0
    values = [v for _, v in pairs]
    spread = max(values) - min(values)
    mean_val = np.mean(values)
    if mean_val == 0:
        return 0.0
    spread_pct = abs(spread / mean_val) * 100
    return min(spread_pct / 30.0, 1.0)


def _score_monotonicity(data_by_value, metric_key="TotalEquivalentWTW"):
    """Score whether metric changes consistently with parameter value (0-1).

    Uses Spearman rank correlation.
    """
    pairs = _get_final_year_values(data_by_value, metric_key)
    if len(pairs) < 3:
        return 0.5  # not enough data to judge
    param_vals = [p for p, _ in pairs]
    metric_vals = [v for _, v in pairs]
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=RuntimeWarning)
        corr, _ = spearmanr(param_vals, metric_vals)
    if np.isnan(corr):
        return 0.5
    return abs(corr)


def _score_cross_scenario_consistency(data_by_value_per_reg, metric_key="TotalEquivalentWTW"):
    """Score whether the same value ranks best across all regulations (0-1).

    Parameters
    ----------
    data_by_value_per_reg : dict
        {regulation: {param_value_str: report_dict}}
    """
    if not data_by_value_per_reg:
        return 0.5

    best_per_reg = []
    for _reg, data_by_value in data_by_value_per_reg.items():
        pairs = _get_final_year_values(data_by_value, metric_key)
        if not pairs:
            continue
        best_val = min(pairs, key=lambda p: p[1])[0]
        best_per_reg.append(best_val)

    if len(best_per_reg) < 2:
        return 0.5

    unique_bests = set(best_per_reg)
    return 1.0 / len(unique_bests)


def _score_fuel_share_realism(data_by_value, regulation):
    """Score how realistic the fuel shares are under no-regulation (0-1).

    Returns 0.5 (neutral) for regulated scenarios.
    """
    if "no_regulation" not in regulation.lower():
        return 0.5

    scores = []
    for _val_str, report in data_by_value.items():
        if report is None or "InstalledPower" not in report:
            continue
        df = report["InstalledPower"]
        row_total = df.iloc[-1].sum()
        if row_total == 0:
            continue

        val_score = 0.0
        n_fuels = 0
        for fuel, (mid, half_w) in _NO_REG_SHARE_BENCHMARKS.items():
            fuel_col = _find_fuel_column(df.columns, fuel)
            if fuel_col is None:
                continue
            actual_share = df.iloc[-1][fuel_col] / row_total
            deviation = abs(actual_share - mid)
            if deviation <= half_w:
                fuel_score = 1.0
            else:
                fuel_score = max(0.0, 1.0 - (deviation - half_w) / 0.20)
            val_score += fuel_score
            n_fuels += 1

        if n_fuels > 0:
            scores.append(val_score / n_fuels)

    return np.mean(scores) if scores else 0.5


def _score_trajectory_smoothness(data_by_value, metric_key="TotalEquivalentWTW"):
    """Score trajectory smoothness (0-1). Penalizes jagged time series."""
    smoothness_scores = []
    for _val_str, report in data_by_value.items():
        if report is None or metric_key not in report:
            continue
        df = report[metric_key]
        if metric_key in SINGLE_METRICS:
            series = df.iloc[:, 0].values.astype(float)
        else:
            series = df.sum(axis=1).values.astype(float)

        if len(series) < 3:
            continue

        amplitude = np.ptp(series)
        if amplitude == 0:
            smoothness_scores.append(1.0)
            continue

        second_deriv = np.diff(series, n=2)
        jerkiness = np.mean(np.abs(second_deriv)) / amplitude
        smoothness_scores.append(max(0.0, 1.0 - jerkiness * 5))

    return np.mean(smoothness_scores) if smoothness_scores else 0.5


# ---------------------------------------------------------------------------
# 3d. Recommendation engine
# ---------------------------------------------------------------------------

def _aggregate_scores(scores_dict):
    """Weighted average of scoring criteria."""
    total = 0.0
    for key, weight in _SCORE_WEIGHTS.items():
        total += scores_dict.get(key, 0.5) * weight
    return total


def _compute_confidence(aggregate_score, sensitivity_score):
    """Determine confidence level from scores."""
    if aggregate_score >= 0.7 and sensitivity_score >= 0.3:
        return "HIGH"
    elif aggregate_score >= 0.5:
        return "MEDIUM"
    return "LOW"


def _build_recommendations(runs, data_by_folder, param_defaults):
    """Build recommendations for each parameter.

    Returns
    -------
    list[Recommendation]
    """
    by_param = defaultdict(list)
    for run in runs:
        by_param[run.param_token].append(run)

    recommendations = []

    for param_token in sorted(by_param.keys()):
        param_runs = by_param[param_token]
        regulations = sorted(set(r.regulation for r in param_runs))
        param_values = sorted(set(r.param_value for r in param_runs), key=_safe_float)

        # Build data_by_value per regulation
        data_by_value_per_reg = {}
        for reg in regulations:
            dbv = {}
            for run in param_runs:
                if run.regulation == reg:
                    dbv[run.param_value] = data_by_folder.get(run.folder_name)
            data_by_value_per_reg[reg] = dbv

        # Score each parameter value
        best_score = -1
        best_value = None
        best_breakdown = {}

        for val_str in param_values:
            # Per-regulation scores averaged
            sens_scores = []
            mono_scores = []
            smooth_scores = []
            realism_scores = []
            for reg in regulations:
                dbv = data_by_value_per_reg[reg]
                sens_scores.append(_score_sensitivity(dbv))
                mono_scores.append(_score_monotonicity(dbv))
                smooth_scores.append(_score_trajectory_smoothness(dbv))
                single_val_dbv = {val_str: dbv.get(val_str)}
                realism_scores.append(_score_fuel_share_realism(single_val_dbv, reg))

            scores = {
                "sensitivity": np.mean(sens_scores),
                "monotonicity": np.mean(mono_scores),
                "cross_scenario": _score_cross_scenario_consistency(data_by_value_per_reg),
                "fuel_share_realism": np.mean(realism_scores),
                "smoothness": np.mean(smooth_scores),
            }

            agg = _aggregate_scores(scores)
            if agg > best_score:
                best_score = agg
                best_value = val_str
                best_breakdown = {**scores, "aggregate": agg}

        confidence = _compute_confidence(best_score, best_breakdown.get("sensitivity", 0))

        recommendations.append(Recommendation(
            param_token=param_token,
            recommended_value=best_value,
            confidence=confidence,
            scores_breakdown=best_breakdown,
            current_default=param_defaults.get(param_token),
        ))

    return recommendations


# ---------------------------------------------------------------------------
# 3e. Round 2 discretization suggestions
# ---------------------------------------------------------------------------

_ROUND2_GRIDS = {
    "fleet_inertia": [0.0, 0.25, 0.5, 0.6, 0.7, 0.75, 0.8, 0.9],
    "bunkering_inertia": [0.0, 0.25, 0.5, 0.6, 0.7, 0.8, 0.9],
    "fleet_beta": [0.5, 1.0, 1.5, 2.0, 3.0, 4.0, 5.0],
    "producer_inertia": [0.0, 0.5, 0.7, 0.8, 0.9],
    "producer_inter_beta": [1, 3, 5],
    "producer_intra_beta": [1, 3, 5],
}


def _suggest_next_discretization(recommendations):
    """Suggest Round 2 parameter grids based on Round 1 findings.

    Returns
    -------
    dict[str, list[float]]
        Suggested parameter values for Round 2.
    """
    result = {}
    for rec in recommendations:
        pt = rec.param_token
        if pt in _ROUND2_GRIDS:
            result[pt] = _ROUND2_GRIDS[pt]
        else:
            v = _safe_float(rec.recommended_value)
            result[pt] = sorted(set([v * 0.5, v * 0.75, v, v * 1.25, v * 1.5]))
    return result


# ---------------------------------------------------------------------------
# 4. Build summary table
# ---------------------------------------------------------------------------

def _build_summary_tables(runs, data_by_folder, param_defaults):
    """Build summary heatmap data: final-year metric values per param value.

    Returns
    -------
    dict
        {regulation: pd.DataFrame} where rows = param_token, cols = param_value,
        values = final-year TotalEquivalentWTW.
    """
    groups = defaultdict(list)
    for run in runs:
        groups[run.regulation].append(run)

    summaries = {}

    for regulation, reg_runs in sorted(groups.items()):
        # Group by param_token
        by_param = defaultdict(list)
        for run in reg_runs:
            by_param[run.param_token].append(run)

        rows = []
        for param_token in sorted(by_param.keys()):
            param_runs = sorted(by_param[param_token], key=lambda r: _safe_float(r.param_value))
            row = {"parameter": param_token}
            for run in param_runs:
                report = data_by_folder.get(run.folder_name)
                if report and "TotalEquivalentWTW" in report:
                    df = report["TotalEquivalentWTW"]
                    final_val = df.iloc[-1, 0]
                    is_def = str(run.param_value) == str(param_defaults.get(param_token))
                    label = f"{run.param_value}{'*' if is_def else ''}"
                    row[label] = final_val
            rows.append(row)

        if rows:
            summaries[regulation] = pd.DataFrame(rows).set_index("parameter")

    return summaries


# ---------------------------------------------------------------------------
# 5. Write HTML dashboard
# ---------------------------------------------------------------------------

def _write_dashboard(figures, summaries, output_path,
                     fuel_share_figures=None, recommendations=None,
                     round2_grids=None):
    """Assemble a single interactive HTML dashboard.

    Parameters
    ----------
    figures : dict
        {(regulation, param_token): list[go.Figure]}
    summaries : dict
        {regulation: pd.DataFrame}
    output_path : str
        Path for the output HTML file.
    fuel_share_figures : dict, optional
        {(regulation, param_token): list[go.Figure]}
    recommendations : list[Recommendation], optional
    round2_grids : dict[str, list[float]], optional
    """
    import plotly.io as pio

    # Gather regulation scenarios
    regulations = sorted(set(reg for reg, _ in figures.keys()))
    if not regulations:
        logger.error("No figures to write — no completed simulations found.")
        return

    # Gather param tokens
    param_tokens = sorted(set(pt for _, pt in figures.keys()))

    fuel_share_figures = fuel_share_figures or {}

    # Start building HTML
    parts = []
    parts.append(_HTML_HEAD)

    # Tab buttons — Recommendations first if available
    parts.append('<div class="tabs">')
    tab_ids = []
    if recommendations:
        tab_ids.append(("recommendations", "Recommendations"))
    for reg in regulations:
        tab_ids.append((reg, reg.replace("_", " ").title()))

    for i, (tid, label) in enumerate(tab_ids):
        active = ' class="active"' if i == 0 else ""
        parts.append(f'  <button{active} onclick="showTab(\'{tid}\')" id="btn-{tid}">{label}</button>')
    parts.append("</div>")

    # Recommendations tab
    if recommendations:
        is_first = True
        display = "block" if is_first else "none"
        parts.append(f'<div class="tab-content" id="tab-recommendations" style="display:{display}">')
        parts.append("<h2>Parameter Recommendations</h2>")
        parts.append(_render_recommendations_html(recommendations, round2_grids))
        parts.append("</div>")

    # Regulation tabs
    for i, reg in enumerate(regulations):
        tab_idx = i + (1 if recommendations else 0)
        display = "block" if tab_idx == 0 else "none"
        parts.append(f'<div class="tab-content" id="tab-{reg}" style="display:{display}">')
        parts.append(f"<h2>{reg.replace('_', ' ').title()}</h2>")

        for pt in param_tokens:
            key = (reg, pt)
            if key not in figures:
                continue

            parts.append('<div class="param-section">')
            parts.append(f"<h3>{pt.replace('_', ' ').title()}</h3>")
            parts.append('<div class="plot-grid">')

            for fig in figures[key]:
                html_div = pio.to_html(fig, full_html=False, include_plotlyjs=False)
                parts.append(f'<div class="plot-cell">{html_div}</div>')

            parts.append("</div>")  # plot-grid

            # Fuel share plots row
            if key in fuel_share_figures:
                parts.append('<h4 style="margin-top:12px;color:#3c5e86;">Fuel Share Breakdown</h4>')
                parts.append('<div class="plot-grid">')
                for fig in fuel_share_figures[key]:
                    html_div = pio.to_html(fig, full_html=False, include_plotlyjs=False)
                    parts.append(f'<div class="plot-cell">{html_div}</div>')
                parts.append("</div>")

            parts.append("</div>")  # param-section

        # Summary table
        if reg in summaries:
            parts.append('<div class="param-section">')
            parts.append("<h3>Summary: Final-Year WTW Emissions</h3>")
            parts.append(_render_summary_table(summaries[reg]))
            parts.append("</div>")

        parts.append("</div>")  # tab-content

    parts.append(_HTML_TAIL)

    html = "\n".join(parts)
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with open(output_path, "w") as f:
        f.write(html)

    logger.info("Dashboard written to %s", output_path)


def _render_recommendations_html(recommendations, round2_grids=None):
    """Render the Recommendations tab content as HTML."""
    _CONF_STYLES = {
        "HIGH": ("background:#eefae8;border:2px solid #286464;color:#23464b", "HIGH"),
        "MEDIUM": ("background:#fcf8e4;border:2px solid #a2703c;color:#a2703c", "MEDIUM"),
        "LOW": ("background:#ffeeea;border:2px solid #804040;color:#804040", "LOW"),
    }

    parts = []

    # Parameter cards
    parts.append(
        '<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(380px,1fr));gap:16px;margin-bottom:24px;">'
    )
    for rec in recommendations:
        style, label = _CONF_STYLES.get(rec.confidence, _CONF_STYLES["LOW"])
        bd = rec.scores_breakdown

        # Score bar: simple inline SVG bars
        bar_html = _render_score_bars(bd)

        parts.append(f'''<div class="param-section" style="margin-bottom:0;">
  <h3 style="margin-bottom:8px;">{rec.param_token.replace('_', ' ').title()}</h3>
  <div style="display:flex;align-items:center;gap:12px;margin-bottom:10px;">
    <span style="font-size:28px;font-weight:700;color:#2c4068;">{rec.recommended_value}</span>
    <span style="padding:4px 12px;border-radius:12px;font-size:13px;font-weight:600;{style}">{label}</span>
    <span style="color:#888;font-size:13px;">default: {rec.current_default}</span>
  </div>
  <div style="font-size:13px;color:#585858;margin-bottom:6px;">
    Aggregate: <strong>{bd.get("aggregate", 0):.2f}</strong>
  </div>
  {bar_html}
  {('<div style="margin-top:8px;padding:6px 10px;background:#ffeeea;border-radius:6px;font-size:12px;color:#804040;">'
    'Needs human review — low confidence in automated recommendation</div>')  # noqa: E122
    if rec.confidence == "LOW" else ""}  # noqa: E122
</div>''')

    parts.append("</div>")  # grid

    # Round 2 discretization table
    if round2_grids:
        parts.append('<div class="param-section">')
        parts.append("<h3>Suggested Round 2 Discretization</h3>")
        parts.append('<table style="border-collapse:collapse;width:100%;font-size:14px;">')
        parts.append('<tr style="background:#2c4068;color:white;">'
                     '<th style="padding:8px 12px;text-align:left;">Parameter</th>'
                     '<th style="padding:8px 12px;text-align:left;">Values</th>'
                     '<th style="padding:8px 12px;text-align:center;">Count</th>'
                     '<th style="padding:8px 12px;text-align:left;">Rationale</th></tr>')

        _RATIONALES = {
            "fleet_inertia": "Coarse below 0.5, fine grid 0.5\u20130.9 where behavior changes",
            "bunkering_inertia": "Coarse low end, dense 0.5\u20130.9",
            "fleet_beta": "Full range; 0.5 kept as outlier reference",
            "producer_inertia": "Coarse low end, denser toward 1",
            "producer_inter_beta": "Coarse grid confirms insensitivity",
            "producer_intra_beta": "Coarse grid confirms insensitivity",
        }

        total_sims = 0
        for pt, vals in sorted(round2_grids.items()):
            vals_str = ", ".join(str(v) for v in vals)
            n = len(vals)
            total_sims += n
            rationale = _RATIONALES.get(pt, "")
            parts.append(f'<tr style="border-bottom:1px solid #dcdcdc;">'
                         f'<td style="padding:8px 12px;font-weight:500;">{pt}</td>'
                         f'<td style="padding:8px 12px;">{vals_str}</td>'
                         f'<td style="padding:8px 12px;text-align:center;">{n}</td>'
                         f'<td style="padding:8px 12px;color:#585858;">{rationale}</td></tr>')

        n_scenarios = 3
        parts.append(f'<tr style="background:#f2f2f2;font-weight:600;">'
                     f'<td style="padding:8px 12px;" colspan="2">Total simulations '
                     f'({total_sims} values \u00d7 {n_scenarios} scenarios)</td>'
                     f'<td style="padding:8px 12px;text-align:center;">{total_sims * n_scenarios}</td>'
                     f'<td></td></tr>')

        parts.append("</table>")
        parts.append("</div>")

    return "\n".join(parts)


def _render_score_bars(scores_breakdown):
    """Render horizontal score bars as inline HTML."""
    _SCORE_COLORS = {
        "sensitivity": "#2c4068",
        "monotonicity": "#3c5e86",
        "cross_scenario": "#68a4c2",
        "fuel_share_realism": "#286464",
        "smoothness": "#96c8e4",
    }
    _SCORE_LABELS = {
        "sensitivity": "Sensitivity",
        "monotonicity": "Monotonicity",
        "cross_scenario": "Cross-scenario",
        "fuel_share_realism": "Fuel share realism",
        "smoothness": "Smoothness",
    }

    parts = ['<div style="font-size:12px;">']
    for key in _SCORE_WEIGHTS:
        val = scores_breakdown.get(key, 0)
        color = _SCORE_COLORS.get(key, "#888")
        label = _SCORE_LABELS.get(key, key)
        pct = max(0, min(100, val * 100))
        parts.append(
            f'<div style="display:flex;align-items:center;gap:6px;margin-bottom:3px;">'
            f'<span style="width:110px;text-align:right;color:#585858;">{label}</span>'
            f'<div style="flex:1;background:#f2f2f2;border-radius:3px;height:14px;">'
            f'<div style="width:{pct:.0f}%;background:{color};height:100%;border-radius:3px;"></div>'
            f'</div>'
            f'<span style="width:36px;color:#585858;">{val:.2f}</span>'
            f'</div>'
        )
    parts.append('</div>')
    return "\n".join(parts)


def _render_summary_table(summary_data):
    """Render per-parameter summary as an HTML table with colored cells.

    Parameters
    ----------
    summary_data : dict
        {param_token: list[dict]} where each dict has 'label', 'value', 'is_default'.
        Or pd.DataFrame with param_token as index.

    Returns colored HTML table where each parameter row has its own values.
    """
    # summary_data is a DataFrame: rows = param_token, cols = "{value}" or "{value}*"
    # But columns differ per row (NaN for non-applicable). Restructure.
    parts = ['<table style="border-collapse:collapse;width:100%;font-size:14px;">']
    parts.append('<tr style="background:#2c4068;color:white;">'
                 '<th style="padding:8px 16px;text-align:left;">Parameter</th>'
                 '<th style="padding:8px 16px;text-align:left;">Values (final-year WTW emissions)</th>'
                 '</tr>')

    for param_token in summary_data.index:
        row = summary_data.loc[param_token].dropna()
        if row.empty:
            continue

        values = [(col, float(val)) for col, val in row.items()]
        if not values:
            continue

        vals_only = [v for _, v in values]
        v_min, v_max = min(vals_only), max(vals_only)
        v_range = v_max - v_min if v_max > v_min else 1.0

        cells_html = []
        for label, val in values:
            # Normalize 0 (best/green) to 1 (worst/red)
            norm = (val - v_min) / v_range if v_range > 0 else 0.5
            # Interpolate green -> yellow -> red
            if norm < 0.5:
                r = int(238 + (250 - 238) * norm * 2)
                g = int(250 + (230 - 250) * norm * 2)
                b = int(232 + (170 - 232) * norm * 2)
            else:
                t = (norm - 0.5) * 2
                r = int(250 + (128 - 250) * t)
                g = int(230 + (64 - 230) * t)
                b = int(170 + (64 - 170) * t)
            bg = f"rgb({r},{g},{b})"

            fmt_val = f"{val:.2e}" if abs(val) >= 1e6 else f"{val:.1f}"
            is_default = label.endswith("*")
            border = "3px solid #2c4068" if is_default else "1px solid #dcdcdc"
            font_weight = "700" if is_default else "400"

            cells_html.append(
                f'<span style="display:inline-block;padding:6px 12px;margin:2px;'
                f'border-radius:4px;background:{bg};border:{border};'
                f'font-weight:{font_weight};font-size:12px;">'
                f'{label}<br>{fmt_val}</span>'
            )

        parts.append(
            f'<tr style="border-bottom:1px solid #dcdcdc;">'
            f'<td style="padding:8px 16px;font-weight:500;vertical-align:top;">{param_token}</td>'
            f'<td style="padding:8px 12px;">{"".join(cells_html)}</td>'
            f'</tr>'
        )

    parts.append('</table>')
    parts.append('<div style="font-size:11px;color:#888;margin-top:6px;">'
                 'Green = lowest emissions, Red = highest. Bold border = default value.</div>')
    return "\n".join(parts)


_HTML_HEAD = """\
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Calibration Analysis Dashboard</title>
<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
<style>
  body {
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
    margin: 0; padding: 20px;
    background: #fafafa; color: #323232;
  }
  h1 { text-align: center; margin-bottom: 10px; }
  .tabs {
    display: flex; gap: 4px; margin-bottom: 20px;
    border-bottom: 2px solid #dcdcdc; padding-bottom: 0;
  }
  .tabs button {
    padding: 10px 24px; border: none; background: #f2f2f2;
    cursor: pointer; font-size: 14px; border-radius: 6px 6px 0 0;
    color: #585858; font-weight: 500;
  }
  .tabs button.active {
    background: #2c4068; color: white;
  }
  .tabs button:hover:not(.active) {
    background: #dcdcdc;
  }
  .param-section {
    background: white; border-radius: 8px; padding: 16px;
    margin-bottom: 20px; box-shadow: 0 1px 3px rgba(0,0,0,0.1);
  }
  .param-section h3 {
    margin-top: 0; color: #2c4068; border-bottom: 1px solid #d4eef4;
    padding-bottom: 6px;
  }
  .plot-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(450px, 1fr));
    gap: 12px;
  }
  .plot-cell { min-width: 0; }
</style>
</head>
<body>
<h1>Calibration Analysis Dashboard</h1>
"""

_HTML_TAIL = """\
<script>
function showTab(reg) {
  document.querySelectorAll('.tab-content').forEach(el => el.style.display = 'none');
  document.querySelectorAll('.tabs button').forEach(btn => btn.classList.remove('active'));
  document.getElementById('tab-' + reg).style.display = 'block';
  document.getElementById('btn-' + reg).classList.add('active');
  // Trigger Plotly relayout for proper sizing
  window.dispatchEvent(new Event('resize'));
}
</script>
</body>
</html>
"""


# ---------------------------------------------------------------------------
# 6. Public entry point
# ---------------------------------------------------------------------------

def calibration_plot(calibration_dir, output_path=None):
    """Generate a calibration comparison dashboard.

    Parameters
    ----------
    calibration_dir : str
        Directory containing calibration output folders and sampled_parameters.csv.
    output_path : str, optional
        Path for the output HTML file.
        Default: ``{calibration_dir}/calibration_analysis.html``.
    """
    calibration_dir = os.path.abspath(calibration_dir)
    if output_path is None:
        output_path = os.path.join(calibration_dir, "calibration_analysis.html")
    else:
        output_path = os.path.abspath(output_path)

    logger.info("Starting calibration analysis for %s", calibration_dir)

    # 1. Load index
    runs, param_defaults = _load_index(calibration_dir)

    # 2. Read reports (concurrent I/O for faster Excel loading)
    data_by_folder = {}
    missing = 0
    folders_to_read = []
    for run in runs:
        folder_path = os.path.join(calibration_dir, run.folder_name)
        if not os.path.exists(folder_path):
            missing += 1
            continue
        folders_to_read.append((run.folder_name, folder_path))

    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        future_to_name = {
            executor.submit(_read_report, fp): fn
            for fn, fp in folders_to_read
        }
        for future in concurrent.futures.as_completed(future_to_name):
            folder_name = future_to_name[future]
            report = future.result()
            if report is not None:
                data_by_folder[folder_name] = report

    logger.info("Read %d reports (%d folders missing, %d runs total)",
                len(data_by_folder), missing, len(runs))

    if not data_by_folder:
        logger.error("No simulation reports found. Check that simulations have completed.")
        return

    # 3. Build figures
    figures = _build_figures(runs, data_by_folder, param_defaults)

    # 3b. Build fuel share figures
    fuel_share_figures = _build_fuel_share_figures(runs, data_by_folder, param_defaults)

    # 4. Build summary tables
    summaries = _build_summary_tables(runs, data_by_folder, param_defaults)

    # 5. Build recommendations
    recommendations = _build_recommendations(runs, data_by_folder, param_defaults)
    round2_grids = _suggest_next_discretization(recommendations)

    for rec in recommendations:
        logger.info("Recommendation: %s = %s (confidence: %s, aggregate: %.2f)",
                    rec.param_token, rec.recommended_value, rec.confidence,
                    rec.scores_breakdown.get("aggregate", 0))

    # 6. Write dashboard
    _write_dashboard(figures, summaries, output_path,
                     fuel_share_figures=fuel_share_figures,
                     recommendations=recommendations,
                     round2_grids=round2_grids)
